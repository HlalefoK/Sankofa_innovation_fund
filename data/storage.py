import openpyxl
from flask import Flask, render_template, request, redirect

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('./templates/application.html')

@app.route('/submit', methods=['POST'])
def submit_form():
    if request.method == 'POST':
        # Get form data
        category = request.form['category']
        business_stage = request.form['business_stage']
        years = int(request.form['years'])
        revenue = int(request.form['revenue'])
        email = request.form['email']
        company_name = request.form['company_name']
        contact = request.form['contact']
        team = int(request.form['team'])
        description = request.form['description']
        funding = int(request.form['funding'])

        # Create or load an existing Excel workbook
        workbook = openpyxl.load_workbook('form_data.xlsx')

        # Select the worksheet where you want to store the data (create if it doesn't exist)
        if 'Form Data' in workbook.sheetnames:
            sheet = workbook['Form Data']
        else:
            sheet = workbook.create_sheet(title='Form Data')

            # Add headers if the sheet is newly created
            headers = ['Category', 'Business Stage', 'Years', 'Revenue', 'Email', 'Company Name', 'Contact', 'Team', 'Description', 'Funding']
            sheet.append(headers)

        # Append form data to the Excel sheet
        data = [category, business_stage, years, revenue, email, company_name, contact, team, description, funding]
        sheet.append(data)

        # Save the workbook
        workbook.save('form_data.xlsx')

        # Redirect to a thank you page
        return redirect('https://hlalefok.github.io/Sankofa_innovation_fund/thanks.html')

if __name__ == '__main__':
    app.run(debug=True)
